# Code to open an excel workbook and read values and do X and Y line of best fit formula
# Modification is needed if the raw data does not get processed with at least x = sqrt() or y = i/sqrt(x)
# This is for 7 columns
# By: John Marcial Email - jmarc114@fiu.edu, Phone - (305)-510-7578
# Contact if shit goes wrong or if you don't understand how to use the program

import math
import openpyxl as xl
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

wb = xl.load_workbook('DD3.xlsx')  # opening workbook of excel sheet
ws = wb['I calculated K1']  # opening sheet in workbook

# all x values for k1 formula graph
v1 = math.sqrt(0.2)
v2 = math.sqrt(0.18)
v3 = math.sqrt(0.15)
v4 = math.sqrt(0.12)
v5 = math.sqrt(0.10)
v6 = math.sqrt(0.08)
v7 = math.sqrt(0.05)

# creating x points of graph grid
x_Table_Var = [v1, v2, v3, v4, v5, v6, v7]

n = len(x_Table_Var)  # number of x points for n in the k1 line of best fit formula (Constant)
xSum = sum(x_Table_Var)  # sum total of x values for k1 line of best fit formula (Constant)

# Warning: Remember to save whatever workbook you are on if you placed any new values otherwise you'll get errors
# obtaining spreadsheet cell values of a certain range
xl_data = []
for row in range(3, 5059):  # (x, y) x for starting position on excel row and y for final + 1
    for col in range(2, 9):  # (x, y) x for starting position on excel column (A,B,C) and y for final + 1
        char = get_column_letter(col)
        xl_data.append(ws[char + str(row)].value)
# print(xl_data)  # (Visual aid only)
# print(len(xl_data))  # (visual confirmation only)
# print(type(xl_data))  # (visual confirmation only)

# making a list of obtained spreadsheet data to create multiple y grid points
splits = np.array_split(xl_data, len(xl_data)/7)

# y grid point finder for k1 formula (MATH)
xl_Data_Point = []
for array in splits:
    xl_Data_Point.append(array / x_Table_Var)
# print(xl_Data_Point)  # (Visual aid only)
# print(type(xl_Data_Point))  # (visual confirmation only)
# print(type(splits))  # (visual confirmation only)

# The product of all (x * y) combinations for different k1s (Changing)
splits2 = np.array_split(xl_Data_Point, len(xl_data)/7)
xy_product_list = []
for array4 in splits2:
    xy_product_list.append(array4 * x_Table_Var)
# print(xy_product_list) # (visual aid only)

# changing the array style list to pure list
xy_sum = []
for array5 in xy_product_list:
    xy_sum.append(sum(array5))
# print(xy_sum) # (visual aid only)

# xy_sum of (x * y) product for k1 changing formula (OK)
xy_sum_real = []
for array6 in xy_sum:
    xy_sum_real.append(sum(array6))
# print(xy_sum_real) # (visual aid only)

# formatting the y grid points in 1x5 lists (VISUAL)
y_Math_Formatting = []
for array2 in xl_Data_Point:
    y_Math_Formatting.append(array2)
    # print(list(array2)) # (Visual aid only)
# print(y_Math_Formatting) # (Visual aid only)

# sum of all y points for one trial of Voltage (Sum of y points for k1 formula) (OK)
y_sum = []
for array3 in y_Math_Formatting:
    y_sum.append(sum(array3))
# print(y_sum)  # (Visual aid only)

# sum of x^2 (Constant) (OK)
xx_Multi = []
for i1, i2 in zip(x_Table_Var, x_Table_Var):
    xx_Multi.append(i1 * i2)
xx_Multi_Sum = sum(xx_Multi) * n
# print(xx_Multi_Sum)  # (visual aid only)

# the product of the sum of x and sum of y (OK)
splits3 = np.array_split(y_sum, len(xl_data)/7)
x_sum_y_sum = []
for array5 in splits3:
    x_sum_y_sum.append(array5 * xSum)
# print(x_sum_y_sum)  # (visual aid only)

# the k1 formula (FUCK YEAH BITCH IM A GOD)
k1_list_raw = []
for i4, i5 in zip(xy_sum_real, x_sum_y_sum):
    k1_list_raw.append(((n * i4)-i5) / (xx_Multi_Sum - (xSum * xSum)))
# print(k1_list_raw)  # (visual aid only)
# print(len(k1_list_raw))  # (visual confirmation only)

# Taking out the array word out of the raw k1 list package and transforming that into a list
k1_processed = []
for arr in k1_list_raw:
    k1_processed.append(sum(arr))
# print(k1_processed)  # (visual aid only)
# print(len(k1_processed))  # (visual confirmation only)

# printing the processed k1 values on a separate work book in excel
# wb2 = Workbook()  # delete 1st (#) when using
# ws2 = wb2.active  # delete 1st (#) when using
# ws2.title = "Data"  # delete 1st (#) when using

# Saving the data to a new excel book and sheet
# ws2.append(k1_processed)  # delete 1st (#) when using
# wb2.save("")  # delete 1st (#) when using
